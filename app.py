import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- PAGE CONFIG ---
st.set_page_config(page_title="VW RTO Verifier", layout="wide")

# --- HELPER FUNCTIONS ---

def normalize_text(text):
    if not text: return ""
    text = re.sub(r'[^\w\s]', ' ', str(text))
    return text.lower().strip()

def check_name_match(excel_name, doc_name):
    # Safe check: if excel_name is a Series (duplicate col), take the first one
    if isinstance(excel_name, pd.Series):
        excel_name = str(excel_name.iloc[0])
    
    if not doc_name or not excel_name:
        return False
    
    clean_excel = normalize_text(excel_name)
    clean_doc = normalize_text(doc_name)

    excel_tokens = clean_excel.split()
    doc_tokens = clean_doc.split()

    matches = 0
    for doc_word in doc_tokens:
        if doc_word in excel_tokens:
            matches += 1
            continue
        if len(doc_word) == 1:
            if any(token.startswith(doc_word) for token in excel_tokens):
                matches += 1
                continue

    if len(doc_tokens) > 0 and (matches / len(doc_tokens)) >= 0.5:
        return True
    return False

def find_column_case_insensitive(columns, allowed_names):
    """
    Strict case-insensitive search.
    Returns the actual column name from the dataframe if found.
    """
    # Normalize allowable names to lower case
    allowed_lower = [name.lower().strip() for name in allowed_names]
    
    # Check each column in the dataframe
    for col in columns:
        if str(col).lower().strip() in allowed_lower:
            return col
    
    return None

def extract_text_from_pdf_upload(uploaded_file):
    text_content = ""
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text: text_content += text + "\n"
    except Exception as e:
        return ""
    return text_content

def parse_document_data(text):
    data = {}
    
    # --- 1. KEYWORD SEARCH (Global Check) ---
    temp_keyword_pattern = r'(temporary\s*registration|temp\s*regn)'
    has_temp_keyword = bool(re.search(temp_keyword_pattern, text, re.IGNORECASE))

    # --- 2. FIND VEHICLE NUMBER ---
    # Standard: MH01AB1234 | BH Series: 22BH1234AA
    perm_pattern = r'\b[A-Z]{2}[0-9]{1,2}[A-Z]{1,3}[0-9]{4}\b'
    bh_pattern = r'\b[0-9]{2}BH[0-9]{4}[A-Z]{1,2}\b'
    
    veh_match = re.search(perm_pattern, text) or re.search(bh_pattern, text)
    
    if veh_match:
        data['vehicle_no'] = veh_match.group(0)
        found_perm_number = True
    else:
        if "new" in text.lower():
            data['vehicle_no'] = "NEW"
        else:
            data['vehicle_no'] = "Not Found"
        found_perm_number = False

    # --- 3. DETERMINE REGISTRATION TYPE ---
    if found_perm_number:
        data['reg_type'] = "Permanent"
    elif data['vehicle_no'] == "NEW":
        data['reg_type'] = "Permanent"
    elif has_temp_keyword:
        data['reg_type'] = "Temporary"
    else:
        data['reg_type'] = "Temporary"

    # --- 4. FIND CHASSIS NO ---
    chassis_match = re.search(r'\b[A-HJ-NPR-Z0-9]{17}\b', text)
    data['doc_chassis'] = chassis_match.group(0) if chassis_match else None

    # --- 5. FIND CUSTOMER NAME ---
    name_match = re.search(r'(?:Received From|Customer Name|Name|Mr\.|Ms\.)[:\s\.]*([A-Za-z\s\.]+)', text, re.IGNORECASE)
    if name_match:
        raw_name = name_match.group(1).strip()
        data['doc_name'] = " ".join(raw_name.split()[:4]) 
    else:
        data['doc_name'] = None

    # --- 6. FIND DATES ---
    numeric_pattern = r'\d{2}[-/]\d{2}[-/]\d{4}'
    text_month_pattern = r'\d{1,2}[-\s][A-Za-z]{3}[-\s]\d{4}'
    date_pattern = f'(?:{numeric_pattern}|{text_month_pattern})'
    
    reg_match = re.search(r'(?:Registration|Regn|Reg\.)\s*Date[:\s]*(' + date_pattern + ')', text, re.IGNORECASE)
    data['reg_date_specific'] = reg_match.group(1) if reg_match else None

    rec_match = re.search(r'Receipt\s*date[:\s]*(' + date_pattern + ')', text, re.IGNORECASE)
    data['receipt_date_specific'] = rec_match.group(1) if rec_match else None

    if not data['reg_date_specific'] and not data['receipt_date_specific']:
        any_date = re.search(date_pattern, text)
        data['fallback_date'] = any_date.group(0) if any_date else None
    else:
        data['fallback_date'] = None

    return data

def analyze_row(row, doc_data, df_docs_all):
    # Safe extractor for Row Series (handles duplicates)
    def get_val(col_name):
        val = row.get(col_name)
        if isinstance(val, pd.Series):
            return val.iloc[0]
        return val

    excel_name = get_val('Customer Name')
    excel_chassis = get_val('Chassis number')

    # 0. CHECK IF PROCESSING FAILED
    if not doc_data.get('doc_chassis'):
        # Secondary Lookup: Name Match / Chassis Mismatch
        if excel_name and not df_docs_all.empty:
            for _, doc_row in df_docs_all.iterrows():
                if check_name_match(excel_name, doc_row['doc_name']):
                    return ("Inconclusive Documentation provided - RTO challan/VAHAN screenshot/Tax paid receipt attached is incorrect", 
                            "Hold", "NAME MATCH / CHASSIS MISMATCH")
        
        return "Please verify manually", "Pending", "NO DOCUMENT FOUND"

    # --- PRIMARY CHECK ---
    chassis_match = True 
    name_is_match = check_name_match(excel_name, doc_data.get('doc_name'))
    is_permanent = doc_data['reg_type'] == "Permanent"

    # 1. APPROVED
    if chassis_match and name_is_match and is_permanent:
        return "Approved", "Approve", "None"

    # 2. TEMP REG
    if chassis_match and name_is_match and not is_permanent:
        return ("Incomplete Documentation provided - RTO challan/VAHAN screenshot/Tax paid receipt is not attached.", 
                "Hold", "TEMP REGISTRATION")

    # 3. NAME MISMATCH
    if chassis_match and not name_is_match:
        return ("Inconclusive Documentation provided - RTO challan/VAHAN screenshot/Tax paid receipt attached is incorrect", 
                "Hold", "NAME MISMATCH")

    return "Please verify manually", "Pending", "UNKNOWN ERROR"

def create_colored_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Verification')
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    header = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    status_col_idx = header.get('RTO status')

    if status_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col_idx)
            val = str(cell.value).strip()
            
            if val == "Approve":
                cell.fill = green_fill
            elif val == "Hold":
                cell.fill = yellow_fill
            elif val == "Pending":
                cell.fill = blue_fill
    
    output_final = io.BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final

# --- STREAMLIT UI ---

st.title("VW RTO Verification")
st.markdown("Output strictly follows the order of the uploaded Excel file.")
st.markdown("---") 

col1, col2 = st.columns(2)

with col1:
    st.header("1. Upload Excel Data")
    uploaded_excel = st.file_uploader("Upload your User Input Excel", type=["xlsx", "xls"])

with col2:
    st.header("2. Upload Documents")
    uploaded_pdfs = st.file_uploader("Upload Document PDFs", type=["pdf"], accept_multiple_files=True)

if st.button("🚀 Run Verification"):
    if uploaded_excel and uploaded_pdfs:
        with st.spinner("Processing Documents..."):
            
            # --- A. PROCESS PDFS ---
            extracted_docs = []
            progress_bar = st.progress(0)
            
            for i, pdf_file in enumerate(uploaded_pdfs):
                text = extract_text_from_pdf_upload(pdf_file)
                doc_info = parse_document_data(text)
                if doc_info['doc_chassis']:
                    extracted_docs.append(doc_info)
                progress_bar.progress((i + 1) / len(uploaded_pdfs))
            
            df_docs = pd.DataFrame(extracted_docs)
            st.success(f"Scanned {len(uploaded_pdfs)} files. Found valid data in {len(df_docs)} files.")

            # --- B. LOAD USER EXCEL ---
            try:
                df_user = pd.read_excel(uploaded_excel)
            except Exception as e:
                st.error("❌ Error reading Excel file.")
                st.stop()
            
            # --- STRICT CASE-INSENSITIVE COLUMN MATCHING ---
            
            # Allowed variations for Chassis
            chassis_variations = ['chassis number', 'vin number']
            chassis_col = find_column_case_insensitive(df_user.columns, chassis_variations)
            
            # Allowed variations for Name
            name_variations = ['customer name']
            name_col = find_column_case_insensitive(df_user.columns, name_variations)

            if not chassis_col or not name_col:
                st.error(f"❌ Column Error. \n\nExpected 'Chassis Number' or 'VIN Number' AND 'Customer Name' (Case Insensitive). \n\nFound columns: {list(df_user.columns)}")
                st.stop()
            
            # --- CLEAN COLUMNS BEFORE MERGE ---
            # 1. Rename to Standard
            rename_map = {chassis_col: 'Chassis number', name_col: 'Customer Name'}
            
            # 2. Drop existing columns with target names if they aren't the source
            for target in ['Chassis number', 'Customer Name']:
                if target in df_user.columns and target not in [chassis_col, name_col]:
                    df_user = df_user.drop(columns=[target])
            
            df_user.rename(columns=rename_map, inplace=True)
            
            # 3. Final safety: remove duplicate columns
            df_user = df_user.loc[:, ~df_user.columns.duplicated()]

            # --- C. MERGE ---
            if not df_docs.empty:
                df_user['Chassis number'] = df_user['Chassis number'].astype(str).str.strip()
                df_docs['doc_chassis'] = df_docs['doc_chassis'].astype(str).str.strip()
                merged_df = pd.merge(df_user, df_docs, left_on='Chassis number', right_on='doc_chassis', how='left')
            else:
                merged_df = df_user.copy()
                merged_df['doc_chassis'] = None

            # --- D. ANALYZE ---
            results = []
            for index, row in merged_df.iterrows():
                
                doc_data = {
                    'doc_name': row.get('doc_name'),
                    'doc_chassis': row.get('doc_chassis'),
                    'reg_type': row.get('reg_type'),
                    'vehicle_no': row.get('vehicle_no', "Not Found")
                }
                
                # --- DATE LOGIC ---
                reg_date = row.get('reg_date_specific')
                receipt_date = row.get('receipt_date_specific')
                fallback_date = row.get('fallback_date')

                if reg_date and str(reg_date).strip():
                    final_reg_date = reg_date
                elif receipt_date and str(receipt_date).strip():
                    final_reg_date = receipt_date
                else:
                    final_reg_date = fallback_date

                remark, status, error_type = analyze_row(row, doc_data, df_docs)
                
                # Create Output Row
                output_row = row.to_dict()
                
                # Handle possible Series
                for k, v in output_row.items():
                    if isinstance(v, pd.Series):
                        output_row[k] = v.iloc[0]

                # Clean up artifacts
                for key in ['doc_name', 'doc_chassis', 'reg_type', 'vehicle_no', 
                           'reg_date_specific', 'receipt_date_specific', 'fallback_date']:
                    if key in output_row: del output_row[key]

                output_row['Verification Date'] = final_reg_date
                output_row['Doc Vehicle Num'] = doc_data['vehicle_no']
                output_row['RTO status'] = status
                output_row['Remarks'] = remark
                
                results.append(output_row)

            final_df = pd.DataFrame(results)

            # Reorder
            cols = list(final_df.columns)
            priority = ['Chassis number', 'Customer Name', 'RTO status', 'Remarks']
            new_order = priority + [c for c in cols if c not in priority]
            final_df = final_df[new_order]

            st.write("### Verification Results")
            st.dataframe(final_df)

            processed_excel = create_colored_excel(final_df)
            
            st.download_button(
                label="📥 Download Colored Excel Report",
                data=processed_excel,
                file_name="verification_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.error("Please upload both the Excel file and the PDF documents.")
